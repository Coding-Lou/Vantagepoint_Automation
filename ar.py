import util
import local_log
import requests
import re
import os
import csv
from tqdm import tqdm
from openpyxl import Workbook
from datetime import datetime
import zipfile

HEADERS = util.set_headers()
# Email
MAIL_FROM = util.get_config(['AR','FROM'])
CC = util.get_config(['AR', 'CC'])
SUBJECT = util.get_config(['AR', 'SUBJECT'])
BODY = util.get_config(['AR', 'BODY'])
OPTIONALMSG = util.get_config(['AR', 'OPTIONALMSG'])

# Output location
ONEDRIVEDIR = util.get_config(['ONEDRIVEDIR'])
WORKDIR = util.get_config(['WORKDIR'])
# Log Conifg
CONSOLE_OUTPUT = local_log.DualOutput("runtime_log.txt")
RECORDS = 1
STATEMENTDATE = None

def format_amount(val):
    return f"{val:,.2f}" if val and val > 0 else ""

def ar_init():
    util.clear_folder("ar_export")
    global STATEMENTDATE
    STATEMENTDATE = input("Statement Date (format 2025-05-01): ")

def init_output():
    global wb
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["ClientID","From","To","CC","Subject","AttachmentName","AttachmentContent","Body", "ClientName"])


def ar_download_csv():
    try:
        url = "https://qcadeltek03.qcasystems.com/vantagepoint/vision/Reporting/Build"
        payload = {"reportPath":"/Standard/AccountsReceivable/AR Statement","reportOptions":{"baseAlternateRowColor":"","baseBottomMargin":0.5,"baseCulture":"default","baseDefaultCurrencyFormat":"###T###T###D##;(###T###T###D##);#","baseDefaultDateFormat":"M/d/yyyy","baseDefaultHTMLFormatting":"Y","baseDefaultNumberFormat":"###T###T###D##;-###T###T###D##;#","baseFont":"Arial","baseFooterText":"[version] - [options]","baseGridTable":"","baseGroupIndent":0.1,"baseHeadingEndDate":"","baseHeadingRowColor":"","baseHeadingStartDate":"","baseHideDocumentMap":"Y","baseHideSingleLineTotals":"Y","baseLeftMargin":0.5,"defaultPage2Top":0,"baseOrientation":"automatic","baseOverrideHeadingDate":"N","basePageHeight":11,"basePageSize":"letter","basePageWidth":8.5,"baseReportName":"AR Statement","baseRightMargin":0.5,"baseShowBorderLines":"N","baseShowFinalTotals":"N","baseShowTotalsOnHeader":"N","baseStartColumnPosition":1,"baseTopMargin":0.5,"baseUnitOfMeasure":"in","baseUseDashpartLayout":"N","baseUseLookupFilterToGrid":"N","ReportGroups":[],"ReportColumns":[{"heading":"Number","width":0.85,"format":"","align":"left","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"Invoice","username":"","customGridColumnSort":""},{"heading":"Date","width":0.8,"format":"M/d/yyyy","align":"left","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"InvoiceDate","username":"","customGridColumnSort":""},{"heading":"Due Date","width":0.8,"format":"M/d/yyyy","align":"left","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"DueDate","username":"","customGridColumnSort":""},{"heading":"Invoiced","width":0.8,"format":"###T###T###D##;(###T###T###D##);#","align":"right","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"OriginalAmt","username":"","customGridColumnSort":""},{"heading":"Balance Due","width":0.8,"format":"###T###T###D##;(###T###T###D##);#","align":"right","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"Balance","username":"","customGridColumnSort":""}],"ReportSections":[],"baseRecordSelection":"","baseShowDetail":"Y","statementType":"C","radioSD":"radioSD3","radioAU":"InvoiceDate","gracePeriod":"30","firmAlign":"Center","showClientMemo":"N","printLongName":"N","printFirmName":"Y","printByLine":"Y","printAddress":"Y","printProjName":"Y","printProjDesc":"Y","printProjNumber":"N","printFooter":"Y","showInvoiceLeadingZeros":"Y","excludeUReceipts":"Y","byLine":"","address1":"#101 6951 72 Street","address2":"Delta, BC","address3":"V4G 0A2","address4":"","HeaderMsg":"","footerMsg":"QCA Systems Ltd.    \nMain Office: #101 6951 72 Street, Delta, BC, V4G 0A2          Phone: (604)-940-0868          Fax: (604)-940-0869\nNorth Shore Office: #201 197 Forester Street, North Vancouver, BC, V7H 0A6\nAll Invoices are due upon receipt","PrintContactFirstName":"Y","PrintContactLastName":"Y","PrintContactMiddleName":"N","PrintContactPreferredName":"N","PrintContactPrefix":"N","PrintContactSuffix":"N","PrintContactTitle":"N","invoiceAddressee":"1","showFooter":"Y","statementSummary":"Y","agingSummary":"Y","clientSelName":"","MarginAndImages":"[{\"ImageID\":\"Our Firm Block\",\"Type\":\"FirmAddress\",\"TopPosition\":0.07,\"LeftPosition\":2.23,\"ColBand\":\"Header\",\"ImageWidth\":0,\"ImageHeight\":0,\"Item\":1,\"Selected\":\"Y\"},{\"ImageID\":\"Client Address\",\"Type\":\"ClientAddress\",\"TopPosition\":1.55,\"LeftPosition\":0,\"ColBand\":\"Header\",\"ImageWidth\":0,\"ImageHeight\":0,\"Item\":2,\"Selected\":\"Y\"},{\"ImageID\":\"Date Block\",\"Type\":\"DateBlock\",\"TopPosition\":1.17,\"LeftPosition\":5.63,\"ColBand\":\"Header\",\"ImageWidth\":0,\"ImageHeight\":0,\"Item\":3,\"Selected\":\"Y\"},{\"ImageID\":\"Statement Label\",\"Type\":\"StatementLabel\",\"TopPosition\":0.03,\"LeftPosition\":0.04,\"ColBand\":\"Header\",\"ImageWidth\":0,\"ImageHeight\":0,\"Item\":4,\"Selected\":\"Y\"}]","ageDays1":"30","ageDays2":"60","ageDays3":"90","ageDays4":"120","ageDays5":"150","statementDate":STATEMENTDATE+"T00:00:00.000","_desc_saveOptionRole":["","","","",""],"saveOptionRole":["ACCOUNTING","ACCOUNTANT","CONTROLLER","CONTROLLER-RO","[CREATOR_USERNAME]"],"baseOriginalFavoriteId":"DA82BA5C8CD94E2CAD6A96814CC49C5C"}}

        response = requests.post(url, headers=HEADERS,json=payload  )
        data = response.json()
        report_path_raw = data["return"]["ReportPath"]
        report_path = report_path_raw.replace(" ", "%20")

        # Step 2: Get Nonce
        url = "https://qcadeltek03.qcasystems.com/vantagepoint/vision/Security/Nonce"
        payload = {}

        response = requests.post(url, headers=HEADERS, json=payload  )
        nonce = response.json()

        # Step 3: Get Viewer
        url = "https://qcadeltek03.qcasystems.com/vantagepoint/reporting/viewer.aspx??&nonce="+nonce+"&reportPath="+report_path+"&allowSchedule=N&reportName=AR%20Statement"

        response = requests.get(url, headers=HEADERS)
        html = response.text

        report_session = re.search(r"ReportSession=([A-Za-z0-9]+)", html)
        control_id     = re.search(r"ControlID=([A-Za-z0-9]+)", html)
        sqlrsReportViewer = re.search(r'_token="([^"]+)"', html)

        # Step 4: SessionKeepALive
        url = "https://qcadeltek03.qcasystems.com/Vantagepoint/Reporting/Reserved.ReportViewerWebControl.axd?OpType=SessionKeepAlive&ControlID="+control_id.group(1)+"&RSProxy=https%3a%2f%2fqcadeltek03.qcasystems.com%2fReportServer"
        response = requests.post(url, headers=HEADERS)

        # Step 5: Download PDF
        fileName = "result.csv"

        url_pdf = ( "https://qcadeltek03.qcasystems.com"
                "/Vantagepoint/Reporting/Reserved.ReportViewerWebControl.axd"
                f"?ReportSession={report_session.group(1)}"
                "&Culture=1033&CultureOverrides=True"
                "&UICulture=2057&UICultureOverrides=True"
                "&ReportStack=1"
                f"&ControlID={control_id.group(1)}"
                "&RSProxy=https%3a%2f%2fqcadeltek03.qcasystems.com%2fReportServer"
                "&OpType=Export"
                f"&FileName={fileName}"
                "&ContentDisposition=OnlyHtmlInline"
                f"&Format=CSV" )

        response = requests.get(url_pdf, headers = HEADERS,stream=True  )

        if response.status_code == 200:
            csvName = os.path.join(ONEDRIVEDIR, WORKDIR, fileName)
            with open(csvName, "wb") as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
            CONSOLE_OUTPUT.tqdm_write("✅ Full statements lists get")
            CONSOLE_OUTPUT.tqdm_write("")
        else:
            CONSOLE_OUTPUT.tqdm_write(f"❌ Error, status code: {response.status_code}")
            CONSOLE_OUTPUT.tqdm_write("Msg:", response.content[:500])
    except Exception as e:
        CONSOLE_OUTPUT.tqdm_write("❌ Error in download the full AR statement list ")

def ar_process():
    clientNames = set()
    with open(os.path.join(ONEDRIVEDIR, WORKDIR, "result.csv"), mode='r', newline='', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        headers = next(reader)
        for row in reader:
            if len(row) >= 3:
                value = row[2]
                if value:
                    clientNames.add(value)
    zipClientName = []
    for i,clientName in enumerate(tqdm(clientNames, desc="Processing progress: ", unit="clientName") , start=1):
        global DUEINVOICE
        DUEINVOICE = ""
        if (not "QCA Systems Ltd." in clientName):
            CONSOLE_OUTPUT.tqdm_write("-----" + clientName + "-----")
            clientID = util.get_clientID(clientName)
            data = ar_review(clientID)
            email = util.get_vendor_email(clientID)
            if email != '': 
                CONSOLE_OUTPUT.tqdm_write("✅ Vendor email: " + email)
            else:
                CONSOLE_OUTPUT.tqdm_write("❌ " + clientName + " mail not found")
            fileName = ar_download_statement_pdf(clientID, clientName)
            if fileName != '': 
                CONSOLE_OUTPUT.tqdm_write("✅ Vendor's statement download")
            else:
                CONSOLE_OUTPUT.tqdm_write("❌ Error")
            pmList = ""
            if ( (data['Age2'] != "" and int(data['Age2']) > 0) or 
                (data['Age3'] != "" and int(data['Age3']) > 0) or 
                (data['Age4'] != "" and int(data['Age4']) > 0) or 
                (data['Age5'] != "" and int(data['Age5']) > 0) ):
                pmList = ar_get_pm_list(clientID)

            tableContent = '<table border="1" width="500" style="border-collapse: collapse"><thead><tr style="text-align: center;"><th>Invoice</th><th>0-30</th><th>31-45</th><th>46-60</th><th>61-90</th><th>90+</th></tr></thead><tbody>'
            tableContent += ar_generate_invoices_table(clientID)
            tableContent += "</tbody></table>"
            ar_create_record(clientID, clientName, email, fileName, pmList, tableContent)
            ar_details(clientID)
            if ar_need_zip(clientID, clientName):
                zipClientName.append(clientName)
            CONSOLE_OUTPUT.tqdm_write(f"{'Client':<30} {'0-30':>12} {'31-45':>12} {'46-60':>12} {'61-90':>12} {'90+':>12}")
            CONSOLE_OUTPUT.tqdm_write("----------------------------------------------------------------------------------------------------------------")
            CONSOLE_OUTPUT.tqdm_write(f"{clientName[:30]:<30} "
                  f"${data['Age1']:>12,.2f} "
                  f"${data['Age2']:>12,.2f} "
                  f"${data['Age3']:>12,.2f} "
                  f"${data['Age4']:>12,.2f} "
                  f"${data['Age5']:>12,.2f} ")
            CONSOLE_OUTPUT.tqdm_write("")
    return zipClientName

def ar_review(clientID):
    try:
        url = "https://qcadeltek03.qcasystems.com/vantagepoint/vision/ARReview/"+ clientID +"?sumColumns=Total%2CAge1%2CAge2%2CAge3%2CAge4%2CAge5%2CTax%2CInterest%2CRetainage%2CRetainers"
        response = requests.get(url, headers = HEADERS)
        data = response.json()
        return data[0]
    except Exception as e:
        CONSOLE_OUTPUT.tqdm_write("❌ Error in get full invoice of vendor " + clientID)

def ar_details(clientID):
    url = "https://qcadeltek03.qcasystems.com/vantagepoint/vision/ARReview/"+ clientID
    response = requests.get(url, headers = HEADERS)
    records = response.json()
    for r in records:
        if r['Total'] < 0:
            continue
        ar_download_proj_invoice_pdf(r['WBS1'], clientID)
        CONSOLE_OUTPUT.tqdm_write('✅ ' + r['WBS1'] + " invoice download")

def ar_download_statement_pdf(clientID, clientName):
    try:
        # Step 2: Build
        url = "https://qcadeltek03.qcasystems.com/vantagepoint/vision/Reporting/Build"
        payload = {"reportPath":"/Standard/AccountsReceivable/AR Statement","reportOptions":{"baseAlternateRowColor":"","baseBottomMargin":0.5,"baseCulture":"default","baseDefaultCurrencyFormat":"###T###T###D##;(###T###T###D##);#","baseDefaultDateFormat":"M/d/yyyy","baseDefaultHTMLFormatting":"Y","baseDefaultNumberFormat":"###T###T###D##;-###T###T###D##;#","baseFont":"Arial","baseFooterText":"[version] - [options]","baseGridTable":"","baseGroupIndent":0.1,"baseHeadingEndDate":"","baseHeadingRowColor":"","baseHeadingStartDate":"","baseHideDocumentMap":"Y","baseHideSingleLineTotals":"Y","baseLeftMargin":0.5,"defaultPage2Top":0,"baseOrientation":"automatic","baseOverrideHeadingDate":"N","basePageHeight":11,"basePageSize":"letter","basePageWidth":8.5,"baseReportName":"AR Statement","baseRightMargin":0.5,"baseShowBorderLines":"N","baseShowFinalTotals":"N","baseShowTotalsOnHeader":"N","baseStartColumnPosition":1,"baseTopMargin":0.5,"baseUnitOfMeasure":"in","baseUseDashpartLayout":"N","baseUseLookupFilterToGrid":"N","ReportGroups":[],"ReportColumns":[{"heading":"Number","width":0.85,"format":"","align":"left","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"Invoice","username":"","customGridColumnSort":""},{"heading":"Date","width":0.8,"format":"M/d/yyyy","align":"left","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"InvoiceDate","username":"","customGridColumnSort":""},{"heading":"Due Date","width":0.8,"format":"M/d/yyyy","align":"left","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"DueDate","username":"","customGridColumnSort":""},{"heading":"Invoiced","width":0.8,"format":"###T###T###D##;(###T###T###D##);#","align":"right","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"OriginalAmt","username":"","customGridColumnSort":""},{"heading":"Balance Due","width":0.8,"format":"###T###T###D##;(###T###T###D##);#","align":"right","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"Balance","username":"","customGridColumnSort":""}],"ReportSections":[],"baseRecordSelection":"","baseShowDetail":"Y","statementType":"C","radioSD":"radioSD3","radioAU":"InvoiceDate","gracePeriod":"30","firmAlign":"Center","showClientMemo":"N","printLongName":"N","printFirmName":"Y","printByLine":"Y","printAddress":"Y","printProjName":"Y","printProjDesc":"Y","printProjNumber":"N","printFooter":"Y","showInvoiceLeadingZeros":"Y","excludeUReceipts":"Y","byLine":"","address1":"#101 6951 72 Street","address2":"Delta, BC","address3":"V4G 0A2","address4":"","HeaderMsg":"","footerMsg":"QCA Systems Ltd.    \nMain Office: #101 6951 72 Street, Delta, BC, V4G 0A2          Phone: (604)-940-0868          Fax: (604)-940-0869\nNorth Shore Office: #201 197 Forester Street, North Vancouver, BC, V7H 0A6\nAll Invoices are due upon receipt","PrintContactFirstName":"Y","PrintContactLastName":"Y","PrintContactMiddleName":"N","PrintContactPreferredName":"N","PrintContactPrefix":"N","PrintContactSuffix":"N","PrintContactTitle":"N","invoiceAddressee":"1","showFooter":"Y","statementSummary":"Y","agingSummary":"Y","clientSelName":{"pKey":"","name":"","type":"client","whereClauseSearch":"N","isLegacy":"N","searchOptions":[{"name":"selectedResultIds","value":clientID,"type":"client","seq":1,"searchLevel":0,"valueDescription":""}]},"MarginAndImages":"[{\"ImageID\":\"Our Firm Block\",\"Type\":\"FirmAddress\",\"TopPosition\":0.07,\"LeftPosition\":2.23,\"ColBand\":\"Header\",\"ImageWidth\":0,\"ImageHeight\":0,\"Item\":1,\"Selected\":\"Y\"},{\"ImageID\":\"Client Address\",\"Type\":\"ClientAddress\",\"TopPosition\":1.55,\"LeftPosition\":0,\"ColBand\":\"Header\",\"ImageWidth\":0,\"ImageHeight\":0,\"Item\":2,\"Selected\":\"Y\"},{\"ImageID\":\"Date Block\",\"Type\":\"DateBlock\",\"TopPosition\":1.17,\"LeftPosition\":5.63,\"ColBand\":\"Header\",\"ImageWidth\":0,\"ImageHeight\":0,\"Item\":3,\"Selected\":\"Y\"},{\"ImageID\":\"Statement Label\",\"Type\":\"StatementLabel\",\"TopPosition\":0.03,\"LeftPosition\":0.04,\"ColBand\":\"Header\",\"ImageWidth\":0,\"ImageHeight\":0,\"Item\":4,\"Selected\":\"Y\"}]","ageDays1":"30","ageDays2":"60","ageDays3":"90","ageDays4":"120","ageDays5":"150","statementDate": STATEMENTDATE + "T00:00:00.000","_desc_saveOptionRole":["","","","",""],"saveOptionRole":["ACCOUNTING","ACCOUNTANT","CONTROLLER","CONTROLLER-RO","[CREATOR_USERNAME]"],"baseOriginalFavoriteId":"DA82BA5C8CD94E2CAD6A96814CC49C5C"}}

        response = requests.post(url, headers=HEADERS, json=payload  )
        data = response.json()
        report_path_raw = data["return"]["ReportPath"]
        report_path = report_path_raw.replace(" ", "%20")

        # Step 3: Get Nonce
        url = "https://qcadeltek03.qcasystems.com/vantagepoint/vision/Security/Nonce"
        payload = {}
        response = requests.post(url, headers=HEADERS, json=payload  )
        nonce = response.json()

        # Step 4: Get Viewer
        url = "https://qcadeltek03.qcasystems.com/vantagepoint/reporting/viewer.aspx??&nonce="+nonce+"&reportPath="+report_path+"&allowSchedule=N&reportName=AR%20Statement"

        response = requests.get(url, headers=HEADERS )

        html = response.text

        report_session = re.search(r"ReportSession=([A-Za-z0-9]+)", html)
        control_id     = re.search(r"ControlID=([A-Za-z0-9]+)", html)
        sqlrsReportViewer = re.search(r'_token="([^"]+)"', html)

        fileName = "Statement of account - " + clientName + " as of " + STATEMENTDATE + ".pdf"

        url_pdf = ( "https://qcadeltek03.qcasystems.com"
                "/Vantagepoint/Reporting/Reserved.ReportViewerWebControl.axd"
                f"?ReportSession={report_session.group(1)}"
                "&Culture=1033&CultureOverrides=True"
                "&UICulture=2057&UICultureOverrides=True"
                "&ReportStack=1"
                f"&ControlID={control_id.group(1)}"
                "&RSProxy=https%3a%2f%2fqcadeltek03.qcasystems.com%2fReportServer"
                "&OpType=Export"
                f"&FileName={fileName}"
                "&ContentDisposition=OnlyHtmlInline"
                f"&Format=PDF" )

        response = requests.get(url_pdf, headers = HEADERS,stream=True  )

        if response.status_code == 200:
            pdfName = os.path.join(ONEDRIVEDIR, WORKDIR, "ar_export", fileName)
            with open(pdfName, "wb") as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
            return fileName
        return ""
    except Exception as e:
        CONSOLE_OUTPUT.tqdm_write("❌ Error in download statement of " + clientName)

def ar_generate_invoices_table(clientID):
    url = "https://qcadeltek03.qcasystems.com/vantagepoint/vision/ARReview/"+ clientID
    response = requests.get(url, headers = HEADERS)
    records = response.json()
    message = ""
    global DUEINVOICE
    for r in records:
        if r['Total'] < 0:
            continue
        projectID = r['WBS1']
        projectID = projectID.replace("/","[_$2F_]")
        url = "https://qcadeltek03.qcasystems.com/vantagepoint/vision/ARReview/"+ projectID +"/"+ clientID +"/ARReviewDetail"
        response = requests.get(url, headers=HEADERS)
        invoices = response.json()
        for invoice in invoices:
            if invoice.get("Age1", 0) > 0 or invoice.get("Age2", 0) > 0 or invoice.get("Age3", 0) > 0 or invoice.get("Age4", 0) > 0 or invoice.get("Age5", 0) > 0:
                message += (
                    f"<tr>"
                    f'<td align="center">{invoice.get("InvoiceNumber", "")}</td>'
                    f'<td align="right" style="text-align: right;">{format_amount(invoice.get("Age1", 0))}</td>'
                    f'<td align="right" style="text-align: right; color:red;">{format_amount(invoice.get("Age2", 0))}</td>'
                    f'<td align="right" style="text-align: right; color:red;">{format_amount(invoice.get("Age3", 0))}</td>'
                    f'<td align="right" style="text-align: right; color:red;">{format_amount(invoice.get("Age4", 0))}</td>'
                    f'<td align="right" style="text-align: right; color:red;">{format_amount(invoice.get("Age5", 0))}</td>'
                    f'</tr>'
                )
                if invoice.get("Age2", 0) > 0 or invoice.get("Age3", 0) > 0 or invoice.get("Age4", 0) > 0 or invoice.get("Age5", 0) > 0:
                    DUEINVOICE += invoice.get("InvoiceNumber", "") + ", "

    return message

def ar_download_proj_invoice_pdf(projectID, clientId):
    projectID = projectID.replace("/","[_$2F_]")
    url = "https://qcadeltek03.qcasystems.com/vantagepoint/vision/ARReview/"+ projectID +"/"+ clientId +"/ARReviewDetail"
    response = requests.get(url, headers=HEADERS)
    invoices = response.json()
    for invoice in invoices: 
        clientName = invoice['ClientName']
        clientName = clientName.replace("/"," ")
        try:
            url = "https://qcadeltek03.qcasystems.com/vantagepoint/app/Invoices/GetInvoiceFileInfo?invoiceMainWBS1="+invoice['InvoiceMainWBS1']+"&wbs1="+ projectID.replace("[_$2F_]","%2F") +"&invoiceNumber="+ invoice['InvoiceNumber'] +"&creditMemoRefno=&linkCompany="
            response = requests.get(url, headers=HEADERS)

            url = "https://qcadeltek03.qcasystems.com/vantagepoint/vision/InteractiveDetail/" + projectID + "/InvoiceHistory/"+ invoice['InvoiceNumber'] +"/print?printBackupReport=Y&printSupportDocuments=N&DownloadInvoice=N&creditMemo=&hasDraftInvoice=&applicationId=ARReview"
            response = requests.get(url, headers=HEADERS)

            if response.headers.get('Content-Type') == 'application/pdf' and response.content.startswith(b'%PDF'):
                fileName = invoice['InvoiceNumber'] + "_" + clientName + ".pdf"
                pdfName = os.path.join(ONEDRIVEDIR, WORKDIR, "ar_export", fileName)
                with open(pdfName, "wb") as f:
                    for chunk in response.iter_content(chunk_size=8192):
                        if chunk:
                            f.write(chunk)
                ar_create_record(clientId, "", "", fileName, "", "")
                continue

            data = response.json()

            report_path_raw = data["ReportPath"]
            report_path = report_path_raw.replace(" ", "%20")

            url = "https://qcadeltek03.qcasystems.com/vantagepoint/vision/Security/Nonce"
            payload = {}
            response = requests.post(url, headers=HEADERS, json=payload  )
            nonce = response.json()

             # Step 4: Get Viewer
            url = "https://qcadeltek03.qcasystems.com/vantagepoint/reporting/viewer.aspx??&nonce="+nonce+"&reportPath="+report_path+"&allowSchedule=N&reportName=Invoice&embedded=Y&runtimeParameters%5B0%5D%5BshowBillingBackup%5D=1&runtimeParameters%5B1%5D%5BPreInvoice%5D=N&runtimeParameters%5B2%5D%5BHeaderInvoice%5D="+invoice['InvoiceNumber']+"&runtimeParameters%5B3%5D%5BMainWBS1%5D="+projectID.replace("[_$2F_]","%2F")+"&runtimeParameters%5B4%5D%5BmainWBS1Name%5D=&runtimeParameters%5B5%5D%5BInvoice%5D="+invoice['InvoiceNumber']+"&runtimeParameters%5B6%5D%5BActivePeriod%5D="

            response = requests.post(url, headers=HEADERS)

            html = response.text

            report_session = re.search(r"ReportSession=([A-Za-z0-9]+)", html)
            control_id     = re.search(r"ControlID=([A-Za-z0-9]+)", html)
            sqlrsReportViewer = re.search(r'_token="([^"]+)"', html)

            fileName = invoice['InvoiceNumber'] + "_" + clientName + ".pdf"

            url_pdf = ( "https://qcadeltek03.qcasystems.com"
                    "/Vantagepoint/Reporting/Reserved.ReportViewerWebControl.axd"
                    f"?ReportSession={report_session.group(1)}"
                    "&Culture=1033&CultureOverrides=True"
                    "&UICulture=2057&UICultureOverrides=True"
                    "&ReportStack=1"
                    f"&ControlID={control_id.group(1)}"
                    "&RSProxy=https%3a%2f%2fqcadeltek03.qcasystems.com%2fReportServer"
                    "&OpType=Export"
                    f"&FileName={fileName}"
                    "&ContentDisposition=OnlyHtmlInline"
                    f"&Format=PDF" )

            response = requests.get(url_pdf, headers = HEADERS, stream=True  )

            if response.status_code == 200:
                pdfName = os.path.join(ONEDRIVEDIR, WORKDIR, "ar_export", fileName)
                with open(pdfName, "wb") as f:
                    for chunk in response.iter_content(chunk_size=8192):
                        if chunk:
                            f.write(chunk)
                ar_create_record(clientId, "", "", fileName,"", "")
        except Exception as e:
            CONSOLE_OUTPUT.tqdm_write("❌ Error in download " + invoice['InvoiceNumber'] + " of " + clientName)
            continue

def ar_get_pm_list(clientID):
    url = "https://qcadeltek03.qcasystems.com/vantagepoint/vision/ARReview/" + clientID
    response = requests.get(url, headers=HEADERS)
    projList = response.json()
    list = set()
    pmList = ""
    for proj in projList:
        if proj['Age2'] > 0 or proj['Age3'] > 0 or proj['Age4'] > 0 or proj['Age5'] > 0:
            projName = proj['WBS1']
            projName = projName.replace(" ", "%20")
            projName = projName.replace("/", "%2F")
            url = "https://qcadeltek03.qcasystems.com/vantagepoint/app/project/GetProjectPlan?wbs1="+ projName +"&jtdSearch=N"
            response = requests.get(url, headers=HEADERS)
            data = response.json()
            if data['ProjMgrEmail'] != '': 
                CONSOLE_OUTPUT.tqdm_write("✅ "+ proj['WBS1'] +" is due, PM email: " + data['ProjMgrEmail'])
                if not data['ProjMgrEmail'] in list:
                    list.add(data['ProjMgrEmail'] )
                    pmList += data['ProjMgrEmail'] + ";"
            else:
                CONSOLE_OUTPUT.tqdm_write("❌ " +  proj['WBS1'] + " PM email not found")        
    return pmList

def ar_need_zip(clientID, clientName):
    count = 0
    curPath = os.path.join(ONEDRIVEDIR, WORKDIR, "ar_export")
    for file in os.listdir( curPath ):
        if clientName in file:
            count+=1
    return count > 5

def ar_zipfile(clientID, clientName):
    ws = wb["Sheet1"]
    count = 0
    global RECORDS
    for row in range(ws.max_row, 0, -1):
        if clientID == ws[f'A{row}'].value:
            count += 1
    if count > 5:
        for row in range(ws.max_row, 0, -1):
            if clientID == ws[f'A{row}'].value and ("Statement of account" not in str(ws[f'F{row}'].value or "")):
                ws.delete_rows(row)
                RECORDS = RECORDS - 1
        curPath = os.path.join(ONEDRIVEDIR, WORKDIR, "ar_export")
        zipFileName = "Invoices of " + clientName + ".zip"
        zipFilePath = os.path.join(curPath, zipFileName)
        with zipfile.ZipFile(zipFilePath, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file in os.listdir(curPath):
                if file == zipFileName:
                    continue
                if clientName in file and "Statement of account" not in file:
                    file_path = os.path.join(curPath, file)
                    arcname = os.path.relpath(file_path, curPath)
                    zipf.write(file_path, arcname=file)
        
        ar_create_record(clientID, clientName, "", zipFileName, "", "")
        timestamp = datetime.now().strftime("%Y%m%d")
        excelName = os.path.join(ONEDRIVEDIR, WORKDIR, f"Job_{timestamp}.xlsx")
        wb.save(excelName)

def ar_create_record(clientID, clientName, email, fileName, pmList, tableContent):
    ws = wb.active
    ws.title = "Sheet1"
    global DUEINVOICE
    if DUEINVOICE != "":
        DUEINVOICE = (
            "<p>We kindly remind you that the following invoices are due: "
            "<span style='background-color: yellow; color: red; font-weight: bold;'>"
            + DUEINVOICE +
            "</span> If you already paid this invoice or have any questions, let us know!</p>"
        )
    body = "<html><p>Dear <b>" + clientName + "</b></p><p> Please find the attached file for <b> Statement of account - " + clientName + " as of " + STATEMENTDATE + "</b>.</p>" + tableContent + DUEINVOICE + OPTIONALMSG + BODY
    row = [clientID, MAIL_FROM, email, pmList + CC, SUBJECT + clientName + " as of " + STATEMENTDATE, fileName, WORKDIR+"\\ar_export\\"+fileName, body, clientName]
    ws.append(row)
    global RECORDS
    RECORDS += 1

def ar_main():
    util.check_folder("ar_export")
    ar_init()
    init_output()
    ar_download_csv()
    zipClientName = ar_process()
    
    for clientName in zipClientName:
        clientID = util.get_clientID(clientName)
        ar_zipfile(clientID, clientName)
    util.save_excel(wb, RECORDS)

if __name__ == "__main__":
    ar_main()