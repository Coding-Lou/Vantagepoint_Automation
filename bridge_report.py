import util
import requests
import re
import os
from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl import load_workbook
import csv
import local_log
from collections import defaultdict

global HEADERS
HEADERS = util.set_headers()
global searchOptions
searchOptions = None
# Log Conifg
global CONSOLE_OUTPUT
CONSOLE_OUTPUT = local_log.DualOutput("runtime_log.txt")

def format_amount(val):
    return f"{val:,.2f}" if val and val > 0 else ""

def smart_convert(val):
    if val.strip() == '':
        return None
    
    if val.startswith('(') and val.endswith(')'):
        try:
            numeric_part = val[1:-1].replace(',', '')
            return -float(numeric_part)
        except ValueError:
            return val
        
    try:
        return float(val.replace(',', ''))
    except ValueError:
        return val

def to_float(value):
    try:
        if value is None:
            return 0.0
        return float(value)
    except (ValueError, TypeError):
        return 0.0

def download_invoice_register_ytd():
    try:
        # Step 1: Build
        url = "https://qcadeltek03.qcasystems.com/Vantagepoint/vision/Reporting/Build"
        payload = {"reportPath":"/Standard/AccountsReceivable/Invoice Register","reportOptions":{"baseAlternateRowColor":"","baseBottomMargin":0.5,"baseChart3D":"N","baseChartColumn":"Other","baseChartDivisor":"1","baseChartFontSize":8,"baseChartHeight":3,"baseChartLabelLines":"N","baseChartLabels":"none","baseChartLeft":1,"baseChartLegendPosition":"righttop","baseChartSeriesColumn2":"","baseChartSeriesColumn3":"","baseChartShowPosition":"1","baseChartTitle":"","baseChartTop":0.5,"baseChartType":"none","baseChartWidth":6,"baseChartXTitle":"","baseChartYTitle":"Other","baseCulture":"default","baseDefaultCurrencyFormat":"###T###T###D##;(###T###T###D##);#","baseDefaultDateFormat":"yyyy-MM-dd","baseDefaultHTMLFormatting":"Y","baseDefaultNumberFormat":"###T###T###D##;-###T###T###D##;#","baseFont":"Arial","baseFooterText":"[version] - [options]","baseGridTable":"","baseGroupIndent":0.1,"baseHeadingEndDate":"","baseHeadingRowColor":"","baseHeadingStartDate":"","baseHideDocumentMap":"Y","baseHideSingleLineTotals":"N","baseLeftMargin":0.5,"defaultPage2Top":0,"baseOrientation":"automatic","baseOverrideHeadingDate":"N","basePageHeight":11,"basePageSize":"letter","basePageWidth":8.5,"baseReportName":"Invoice Register","baseRightMargin":0.5,"baseShowBorderLines":"N","baseShowFinalTotals":"N","baseShowTotalsOnHeader":"N","baseStartColumnPosition":0.5,"baseTopMargin":0.5,"baseUnitOfMeasure":"in","baseUseDashpartLayout":"N","baseUseLookupFilterToGrid":"N","ReportGroups":[],"ReportColumns":[{"heading":"Project","width":1,"format":"","align":"left","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"WBS1","username":"","customGridColumnSort":""},{"heading":"Date","width":0.7,"format":"yyyy-MM-dd","align":"left","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"TransDate","username":"","customGridColumnSort":""},{"heading":"Invoice","width":1,"format":"","align":"left","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"InvoiceNumber","username":"","customGridColumnSort":""},{"heading":"Total","width":0.7,"format":"###T###T###D##;(###T###T###D##);#","align":"right","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"TotalAmt","username":"","customGridColumnSort":""},{"heading":"Prof Fees","width":0.7,"format":"###T###T###D##;(###T###T###D##);#","align":"right","sectionName":"","sectionRow":0,"sectionColumn":0,"columnID":"Col1","username":"","customGridColumnSort":""},{"heading":"H/W Sales","width":0.7,"format":"###T###T###D##;(###T###T###D##);#","align":"right","sectionName":"","sectionRow":0,"sectionColumn":0,"columnID":"Col2","username":"","customGridColumnSort":""},{"heading":"S/W Sales","width":0.7,"format":"###T###T###D##;(###T###T###D##);#","align":"right","sectionName":"","sectionRow":0,"sectionColumn":0,"columnID":"Col3","username":"","customGridColumnSort":""},{"heading":"O/S Services","width":0.7,"format":"###T###T###D##;(###T###T###D##);#","align":"right","sectionName":"","sectionRow":0,"sectionColumn":0,"columnID":"Col4","username":"","customGridColumnSort":""},{"heading":"Reimbursable","width":0.7,"format":"###T###T###D##;(###T###T###D##);#","align":"right","sectionName":"","sectionRow":0,"sectionColumn":0,"columnID":"Col5","username":"","customGridColumnSort":""},{"heading":"EHF","width":0.7,"format":"###T###T###D##;(###T###T###D##);#","align":"right","sectionName":"","sectionRow":0,"sectionColumn":0,"columnID":"Col6","username":"","customGridColumnSort":""},{"heading":"Other","width":0.7,"format":"###T###T###D##;(###T###T###D##);#","align":"right","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"Other","username":"","customGridColumnSort":""},{"heading":"Taxes","width":0.7,"format":"###T###T###D##;(###T###T###D##);#","align":"right","sectionName":"","sectionRow":0,"sectionColumn":0,"columnID":"Col7","username":"","customGridColumnSort":""},{"heading":"Net Revenue","width":1,"format":"###T###T###T###D##;(###T###T###T###D##);#","align":"right","sectionName":"","sectionRow":0,"sectionColumn":0,"columnID":"Net Revenue","header1":"Net Revenue","header2":"","detailExpression":"[TotalAmt]-[Col7]","groupExpression":"[TotalAmt]-[Col7]","queryJoin":"","checkSecurity":"N","altHeader1":"","altHeader2":"","queryColumn":"","calculatedColumnType":"ALLFRAMES","username":"DPALACIO","customGridColumnSort":""}],"ReportSections":[],"baseRecordSelection":"","baseShowDetail":"Y","baseLeft1":0,"baseRight1":21,"baseLeft2":0,"baseRight2":8,"baseLeft3":0,"baseRight3":8,"baseSub":"1","rollType":"Project","timeframe":"YTD","radioTF":"radio1","tfCYJ":"Y","clientInfo":"None","InterestCol":"0","txtShowLink":"N","baseSelectionRows":11254,"_desc_saveOptionRole":["","",""],"saveOptionRole":["[CREATOR_USERNAME]","ACCOUNTANT","ACCOUNTING"],"baseOriginalFavoriteId":"385322010f1c4997b596c9017169c40b"}}
        response = requests.post(url, headers=HEADERS, json=payload  )
        data = response.json()
        report_path_raw = data["return"]["ReportPath"]
        report_path = report_path_raw.replace(" ", "%20")

        # Step 2: Get Nonce
        url = "https://qcadeltek03.qcasystems.com/vantagepoint/vision/Security/Nonce"
        payload = {}
        response = requests.post(url, headers=HEADERS, json=payload  )
        nonce = response.json()

        # Step 3: Get Viewer
        url = "https://qcadeltek03.qcasystems.com/vantagepoint/reporting/viewer.aspx?&nonce="+nonce+"&ResetReportViewerOnPreview=Y&reportPath="+report_path+"&allowSchedule=Y&origReportPath=/Standard/AccountsReceivable/Invoice%20Register&reportName=Invoice%20Register"

        # Step 4: Get report session
        response = requests.get(url, headers=HEADERS )
        html = response.text
        report_session = re.search(r"ReportSession=([A-Za-z0-9]+)", html)
        control_id = re.search(r"ControlID=([A-Za-z0-9]+)", html)
        sqlrsReportViewer = re.search(r'_token="([^"]+)"', html)

        if not (report_session and control_id):
            raise RuntimeError("Error")
        
        exportFileName = "Invoice Register.csv"

        # Step 5: Download the csv report
        url = ( "https://qcadeltek03.qcasystems.com"
                "/Vantagepoint/Reporting/Reserved.ReportViewerWebControl.axd"
                f"?ReportSession={report_session.group(1)}"
                "&Culture=1033&CultureOverrides=True"
                "&UICulture=2057&UICultureOverrides=True"
                "&ReportStack=1"
                f"&ControlID={control_id.group(1)}"
                "&RSProxy=https%3a%2f%2fqcadeltek03.qcasystems.com%2fReportServer"
                "&OpType=Export"
                f"&FileName=Invoice+Register"
                "&ContentDisposition=OnlyHtmlInline&Format=CSV" )
        
        response = requests.get(url, headers = HEADERS,stream=True  )
        if response.status_code == 200 and response.headers.get("Content-Type") == "text/csv; charset=utf-8":
            csvName = os.path.join("bridge_report", exportFileName)
            if os.path.exists(csvName):
                os.remove(csvName)
            with open(csvName, "wb") as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
            CONSOLE_OUTPUT.tqdm_write("✅ "+ csvName+" Downloaded")

    except Exception as e:
        CONSOLE_OUTPUT.tqdm_write("⚠️ Failed to download the invoices register")

def labour_download_labour_details(projectId):
    projectName = projectId.replace("/", "_")

    buildUrl = "https://qcadeltek03.qcasystems.com/vantagepoint/vision/Reporting/Build"
    payload = {"reportPath":"/Standard/Project/Labor Detail","reportOptions":{"baseAlternateRowColor":"","baseBottomMargin":0.5,"baseChart3D":"N","baseChartColumn":"regHrs","baseChartDivisor":"1","baseChartFontSize":8,"baseChartHeight":3,"baseChartLabelLines":"N","baseChartLabels":"none","baseChartLeft":1,"baseChartLegendPosition":"righttop","baseChartSeriesColumn2":"","baseChartSeriesColumn3":"","baseChartShowPosition":"1","baseChartTitle":"","baseChartTop":0.5,"baseChartType":"none","baseChartWidth":6,"baseChartXTitle":"Project Number","baseChartYTitle":"Regular Hours","baseCulture":"default","baseDefaultCurrencyFormat":"###T###T###D##;(###T###T###D##);#","baseDefaultDateFormat":"yyyy-MM-dd","baseDefaultHTMLFormatting":"Y","baseDefaultNumberFormat":"###T###T###D##;-###T###T###D##;#","baseFont":"Arial","baseFooterText":"[version] - [options]","baseGridTable":"","baseGroupIndent":0.05,"baseHeadingEndDate":"","baseHeadingRowColor":"","baseHeadingStartDate":"","baseHideDocumentMap":"Y","baseHideSingleLineTotals":"Y","baseLeftMargin":0.5,"defaultPage2Top":0,"baseOrientation":"automatic","baseOverrideHeadingDate":"N","basePageHeight":11,"basePageSize":"letter","basePageWidth":8.5,"baseReportName":"Labour Detail","baseRightMargin":0.5,"baseShowBorderLines":"N","baseShowFinalTotals":"N","baseShowTotalsOnHeader":"N","baseStartColumnPosition":3.5,"baseTopMargin":0.5,"baseUnitOfMeasure":"in","baseUseDashpartLayout":"N","baseUseLookupFilterToGrid":"N","ReportGroups":[{"label":"Project Number","sort":"ASC","color":"000080","subTotal":"N","showHeading":"N","pageHeading":"N","collapseExpand":"D","line":"None","pageBreak":"N","groupID":"projectNumber","customGridColumnSort":"","groupWBSLevel":"1"},{"label":"Task Number","sort":"ASC","color":"228B22","subTotal":"N","showHeading":"N","pageHeading":"N","collapseExpand":"D","line":"None","pageBreak":"N","groupID":"wbs2Number","customGridColumnSort":"","groupWBSLevel":"2"},{"label":"Subtask Number","sort":"ASC","color":"000000","subTotal":"N","showHeading":"N","pageHeading":"N","collapseExpand":"D","line":"None","pageBreak":"N","groupID":"wbs3Number","customGridColumnSort":"","groupWBSLevel":"3"},{"label":"Labour Code Level 1","sort":"ASC","color":"000000","subTotal":"N","showHeading":"N","pageHeading":"N","collapseExpand":"D","line":"None","pageBreak":"N","groupID":"labcd1","customGridColumnSort":"","groupWBSLevel":"1"}],"ReportColumns":[{"heading":"Project","width":1.25,"format":"","align":"left","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"WBS1","username":"","customGridColumnSort":""},{"heading":"Task","width":0.8,"format":"","align":"left","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"WBS2","username":"","customGridColumnSort":""},{"heading":"Subtask","width":0.8,"format":"","align":"left","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"WBS3","username":"","customGridColumnSort":""},{"heading":"Date","width":0.75,"format":"yyyy-MM-dd","align":"left","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"transDate","username":"","customGridColumnSort":""},{"heading":"Regular Hours","width":0.55,"format":"###T###T###D##;-###T###T###D##;#","align":"right","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"regHrs","username":"","customGridColumnSort":""},{"heading":"Regular Billing","width":0.85,"format":"###T###T###D##;(###T###T###D##);#","align":"right","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"regAmt","username":"","customGridColumnSort":""},{"heading":"Overtime Hours","width":0.55,"format":"###T###T###D##;-###T###T###D##;#","align":"right","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"ovtHrs","username":"","customGridColumnSort":""},{"heading":"Overtime Billing","width":0.85,"format":"###T###T###D##;(###T###T###D##);#","align":"right","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"ovtAmt","username":"","customGridColumnSort":""},{"heading":"Ovt-2 Hours","width":0.55,"format":"###T###T###D##;-###T###T###D##;#","align":"right","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"specialOvtHrs","username":"","customGridColumnSort":""},{"heading":"Ovt-2 Billing","width":0.85,"format":"###T###T###D##;(###T###T###D##);#","align":"right","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"specialOvtAmt","username":"","customGridColumnSort":""},{"heading":"Total Hours","width":0.55,"format":"###T###T###D##;-###T###T###D##;#","align":"right","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"totHrs","username":"","customGridColumnSort":""},{"heading":"Total Billing","width":0.85,"format":"###T###T###D##;(###T###T###D##);#","align":"right","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"totAmt","username":"","customGridColumnSort":""},{"heading":"Bill Status","width":0.55,"format":"","align":"left","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"billStatus","username":"","customGridColumnSort":""},{"heading":"Period","width":0.55,"format":"","align":"left","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"period","username":"","customGridColumnSort":""},{"heading":"Post Seq","width":0.55,"format":"","align":"left","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"postSeq","username":"","customGridColumnSort":""},{"heading":"Charge Type","width":0.55,"format":"","align":"left","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"chargeType","username":"","customGridColumnSort":""},{"heading":"Trans Type","width":0.55,"format":"","align":"left","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"transType","username":"","customGridColumnSort":""}],"ReportSections":[],"baseRecordSelection":{"pKey":"","name":"","type":"wbs1","whereClauseSearch":"N","isLegacy":"N","searchOptions":[{"name":"selectedResultIds","value":projectId,"type":"wbs1","seq":1,"searchLevel":0,"valueDescription":""}]},"baseCreateActivity":"N","baseShowDetail":"Y","baseLeft1":0,"baseRight1":21,"baseLeft2":0,"baseRight2":8,"baseLeft3":0,"baseRight3":8,"baseSub":"1","rollType":"Project","CurrentWBSActivityActiveWBS1Only":"N","CurrentWBSActivityActiveWBS2Only":"N","CurrentWBSActivityActiveWBS3Only":"N","CurrentWBSActivityActivityRange":"1","CurrentWBSActivityInclInvoiceActivity":"N","timeframeBackup":"JTD","timeFrame":"radio1","periodType":"3","transDetail":"1","curSelection":"2","showComments":"N","showUnposted":"N","atCost":"2","CurrentWBSActivityCheckLabor":"Y","CurrentWBSActivityCheckExpense":"Y","CurrentWBSActivityUnpostedLabor":"N","employeeWhere":"","emOwnerfl":"","dateStart":"","dateEnd":"","periodStart":"","periodEnd":"","baseSelectionRows":1,"saveOptionRole":"[CREATOR_USERNAME]","baseOriginalFavoriteId":"","LaborPostingRun":[]}}

    response = requests.post(buildUrl, headers=HEADERS, json=payload  )
    data = response.json()
    report_path_raw = data["return"]["ReportPath"]
    report_path = report_path_raw.replace(" ", "%20")

    # Step 3: Get Nonce
    nonceUrl = "https://qcadeltek03.qcasystems.com/vantagepoint/vision/Security/Nonce"
    payload = {}

    response = requests.post(nonceUrl, headers=HEADERS, json=payload  )
    nonce = response.json()

    # Step 4: Get Viewer
    url = "https://qcadeltek03.qcasystems.com/vantagepoint/reporting/viewer.aspx??&nonce="+nonce+"&reportPath="+report_path+"&allowSchedule=Y&reportName=Project%20Labour%20Export"

    response = requests.get(url, headers=HEADERS )

    html = response.text

    report_session = re.search(r"ReportSession=([A-Za-z0-9]+)", html)
    control_id     = re.search(r"ControlID=([A-Za-z0-9]+)", html)
    sqlrsReportViewer = re.search(r'_token="([^"]+)"', html)

    if not (report_session and control_id):
        raise RuntimeError("Error")

    labourFileName = "Labour Detail " + projectName + ".csv"

    url_pdf = ( "https://qcadeltek03.qcasystems.com"
            "/Vantagepoint/Reporting/Reserved.ReportViewerWebControl.axd"
            f"?ReportSession={report_session.group(1)}"
            "&Culture=1033&CultureOverrides=True"
            "&UICulture=2057&UICultureOverrides=True"
            "&ReportStack=1"
            f"&ControlID={control_id.group(1)}"
            "&RSProxy=https%3a%2f%2fqcadeltek03.qcasystems.com%2fReportServer"
            "&OpType=Export"
            f"&FileName=Project+Labour+Export"
            "&ContentDisposition=OnlyHtmlInline&Format=CSV" )

    response = requests.get(url_pdf, headers = HEADERS,  stream=True  )

    if response.status_code == 200 and response.headers.get("Content-Type") == "text/csv; charset=utf-8":
        csvName = os.path.join("bridge_report", labourFileName)
        os.makedirs(os.path.dirname(csvName), exist_ok=True)
        with open(csvName, "wb") as f:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
        print("✅ "+ projectName +" timesheet Downloaded")
    with open(csvName, "r", encoding="utf-8") as f:
        lines = f.readlines()
    
    lines = lines[3:]
    
    with open(csvName, "w", encoding="utf-8") as f:
        f.writelines(lines)

def labour_download_transaction_details(projectId):
    projectName = projectId.replace("/", "_")
    try:
        url = "https://qcadeltek03.qcasystems.com/vantagepoint/vision/Reporting/Build"
        payload = {"reportPath":"/Standard/AccountsReceivable/Invoice Transaction Detail","reportOptions":{"baseAlternateRowColor":"","baseBottomMargin":0.5,"baseChart3D":"N","baseChartColumn":"HoursUnits","baseChartDivisor":"1","baseChartFontSize":8,"baseChartHeight":3,"baseChartLabelLines":"N","baseChartLabels":"none","baseChartLeft":1,"baseChartLegendPosition":"righttop","baseChartSeriesColumn2":"","baseChartSeriesColumn3":"","baseChartShowPosition":"1","baseChartTitle":"","baseChartTop":0.5,"baseChartType":"none","baseChartWidth":6,"baseChartXTitle":"Invoice","baseChartYTitle":"Hours / Units","baseCulture":"default","baseDefaultCurrencyFormat":"###T###T###D##;(###T###T###D##);#","baseDefaultDateFormat":"yyyy-MM-dd","baseDefaultHTMLFormatting":"Y","baseDefaultNumberFormat":"###T###T###D##;-###T###T###D##;#","baseFont":"Arial","baseFooterText":"[version] - [options]","baseGridTable":"","baseGroupIndent":0.1,"baseHeadingEndDate":"","baseHeadingRowColor":"","baseHeadingStartDate":"","baseHideDocumentMap":"Y","baseHideSingleLineTotals":"N","baseLeftMargin":0.5,"defaultPage2Top":0,"baseOrientation":"automatic","baseOverrideHeadingDate":"N","basePageHeight":11,"basePageSize":"letter","basePageWidth":8.5,"baseReportName":"Invoice Transaction Detail","baseRightMargin":0.5,"baseShowBorderLines":"N","baseShowFinalTotals":"N","baseShowTotalsOnHeader":"N","baseStartColumnPosition":0.1,"baseTopMargin":0.5,"baseUnitOfMeasure":"in","baseUseDashpartLayout":"N","baseUseLookupFilterToGrid":"N","ReportGroups":[{"label":"Invoice","sort":"ASC","color":"000000","subTotal":"N","showHeading":"Y","pageHeading":"N","collapseExpand":"E","line":"None","pageBreak":"N","groupID":"Invoice","groupWBSLevel":"1"}],"ReportColumns":[{"heading":"Project","width":1,"format":"","align":"left","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"WBS1","username":"","customGridColumnSort":""},{"heading":"Task","width":0.75,"format":"","align":"left","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"WBS2","username":"","customGridColumnSort":""},{"heading":"Subtask","width":0.75,"format":"","align":"left","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"WBS3","username":"","customGridColumnSort":""},{"heading":"Date","width":0.7,"format":"yyyy-MM-dd","align":"left","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"TransDate","username":"","customGridColumnSort":""},{"heading":"Labour Code /Account","width":1,"format":"","align":"left","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"LaborCodeAccount","username":"","customGridColumnSort":""},{"heading":"Employee/ Reference","width":1,"format":"","align":"left","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"EmployeeReference","username":"","customGridColumnSort":""},{"heading":"Description","width":2.4,"format":"","align":"left","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"Description","username":"","customGridColumnSort":""},{"heading":"Hours/ Units","width":0.7,"format":"###T###T###D##;-###T###T###D##;#","align":"right","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"HoursUnits","username":"","customGridColumnSort":""},{"heading":"Billing Amount","width":0.75,"format":"###T###T###D##;(###T###T###D##);#","align":"right","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"BillingAmount","username":"","customGridColumnSort":""},{"heading":"Billing Status","width":0.6,"format":"","align":"left","sectionName":"Section 1","sectionRow":0,"sectionColumn":1,"columnID":"BillingStatus","username":"","customGridColumnSort":""}],"ReportSections":[],"baseRecordSelection":{"pKey":"","name":"","type":"wbs1","whereClauseSearch":"N","isLegacy":"N","searchOptions":[{"name":"selectedResultIds","value":projectId,"type":"wbs1","seq":1,"searchLevel":0,"valueDescription":""}]},"baseShowDetail":"Y","baseLeft1":0,"baseRight1":21,"baseLeft2":0,"baseRight2":8,"baseLeft3":0,"baseRight3":8,"baseSub":"1","rollType":"Project","timeframe":"JTD","radioTF":"radio1","tfCYJ":"J","PrintInvTotals":"Y","SummarizeByEmp":"N","ShowTSComments":"N","ShowBilled":"Y","ShowWriteOff":"Y","ShowDeleted":"Y","ShowNonBill":"Y","IncludeLabor":"Y","IncludeConsultant":"Y","IncludeExpenses":"Y","baseSelectionRows":1,"saveOptionRole":"[CREATOR_USERNAME]","baseOriginalFavoriteId":""}}

        response = requests.post(url, headers=HEADERS, json=payload  )
        data = response.json()
        report_path_raw = data["return"]["ReportPath"]
        report_path = report_path_raw.replace(" ", "%20")

        # Step 3: Get Nonce
        nonceUrl = "https://qcadeltek03.qcasystems.com/vantagepoint/vision/Security/Nonce"
        payload = {}

        response = requests.post(nonceUrl, headers=HEADERS, json=payload  )
        nonce = response.json()

        # Step 4: Get Viewer
        url = "https://qcadeltek03.qcasystems.com/vantagepoint/reporting/viewer.aspx??&nonce="+nonce+"&reportPath="+report_path+"&allowSchedule=Y&reportName=Invoice%20Transaction%20Detail"

        response = requests.get(url, headers=HEADERS )

        html = response.text

        report_session = re.search(r"ReportSession=([A-Za-z0-9]+)", html)
        control_id     = re.search(r"ControlID=([A-Za-z0-9]+)", html)
        sqlrsReportViewer = re.search(r'_token="([^"]+)"', html)

        if not (report_session and control_id):
            raise RuntimeError("Error")

        # Step 5: Download file
        transactionFileName = "Transaction Detail " + projectName + ".csv"
        url_pdf = ( "https://qcadeltek03.qcasystems.com"
                "/Vantagepoint/Reporting/Reserved.ReportViewerWebControl.axd"
                f"?ReportSession={report_session.group(1)}"
                "&Culture=1033&CultureOverrides=True"
                "&UICulture=2057&UICultureOverrides=True"
                "&ReportStack=1"
                f"&ControlID={control_id.group(1)}"
                "&RSProxy=https%3a%2f%2fqcadeltek03.qcasystems.com%2fReportServer"
                "&OpType=Export"
                f"&FileName=Project+Labour+Export"
                "&ContentDisposition=OnlyHtmlInline&Format=CSV" )

        response = requests.get(url_pdf, headers = HEADERS, stream=True  )

        if response.status_code == 200 and response.headers.get("Content-Type") == "text/csv; charset=utf-8":
            csvName = os.path.join("bridge_report", transactionFileName)
            with open(csvName, "wb") as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
            print("✅ "+ projectName +" transaction details Downloaded")

            with open(csvName, "r", encoding="utf-8") as f:
                lines = f.readlines()
    
            lines = lines[3:]
    
            with open(csvName, "w", encoding="utf-8") as f:
                f.writelines(lines)
    except:
        print("error")

def labour_invoice_match(project):
    #key: {employeeID}#{taskNum}#{subtaskNum}#{date}#{billingStatus}#{chargeHours}
    
    projectName = project.replace("/", "_")
    transactionFileName = "Transaction Detail " + projectName + ".csv"
    invoice_map = defaultdict(list)
    used = []
    with open( os.path.join("bridge_report", transactionFileName), newline='', encoding='utf-8') as csvfile:
        reader = list(csv.reader(csvfile))
        header = reader[0]
        used = [False] * len(reader)
        col_map = {col_name.strip(): idx for idx, col_name in enumerate(header)}
        for index, transactionRow in enumerate(reader):
            try:
                invoice = transactionRow[col_map['groupHeader1_GroupColumn']].split(' ')[1]
                taskNum = transactionRow[col_map['detail_WBS2']]
                subtaskNum = transactionRow[col_map['detail_WBS3']]
                date = transactionRow[col_map['detail_TransDate']]
                employeeID = transactionRow[col_map['detail_EmployeeReference']]
                chargeHours = f"{float(transactionRow[col_map['detail_HoursUnits']]):.2f}"
                billingStatus = transactionRow[col_map['detail_BillingStatus']]
                key = f'{employeeID}#{taskNum}#{subtaskNum}#{date}#{chargeHours}'
                invoice_map[key].append(f'{index}#{invoice}')
            except Exception as e:
                continue

    wb = Workbook()
    ws = wb.active
    ws.title = "Labour Details"
    transfer_map = defaultdict(list)

    labourFileName = "Labour Detail " + projectName + ".csv"
    with open( os.path.join("bridge_report", labourFileName), newline='', encoding='utf-8') as csvfile:
        reader = list(csv.reader(csvfile))

        headers = reader[0]
        col_map = {col_name.strip(): idx for idx, col_name in enumerate(headers)}

        index = 1
        for row_index, labourRow in enumerate(reader[0:], start=1):
            if len(labourRow) < len(col_map) or labourRow[col_map['detail_postSeq']] == 0:
                continue

            for col_index, cell_value in enumerate(labourRow):
                header_value = headers[col_index] if col_index < len(headers) else ""
                if 'footer' not in header_value.lower():
                    clean_value = smart_convert(cell_value)
                    ws.cell(row=index, column=col_index + 1, value=clean_value)
            index += 1
            try:
                taskNum = labourRow[col_map['detail_WBS2']]
                subtaskNum = labourRow[col_map['detail_WBS3']]
                date = labourRow[col_map['detail_transDate']]
                employeeID = labourRow[col_map['detail_Employee']]
                billingStatus = labourRow[col_map['detail_billStatus']].strip()
                hours = float(labourRow[col_map['detail_totHrs']] or 0)
                if project.startswith("Q-NBT"):
                    hours = hours + float(labourRow[col_map['detail_ovtAmt']] or 0) + float(labourRow[col_map['detail_specialOvtHrs']] or 0) * 0.5
                    billingStatus = "F"
                chargeHours = f"{hours:.2f}"

                if (billingStatus != "T"):
                    key = f'{employeeID}#{taskNum}#{subtaskNum}#{date}#{chargeHours}'
                    if key in invoice_map:
                        for i in invoice_map[key]:
                            invoice_row = int(i.split("#")[0])
                            invoice_number = f"{i.split('#')[1]}"
                            if not used[invoice_row]:
                                ws.cell(row = index-1, column=col_map['Comment']+1).value = invoice_number
                                used[invoice_row] = True
                                break
                else:
                    key = f'{employeeID}#{taskNum}#{subtaskNum}#{date}'
                    transfer_map[key] = transfer_map.get(key, 0.0) + hours
            
            except Exception as e:
                continue

        # Billing Status T
        index = 1
        for row_index, labourRow in enumerate(reader[1:], start=1):
            if len(labourRow) < len(col_map) or labourRow[col_map['detail_postSeq']] == 0:
                continue

            for col_index, cell_value in enumerate(labourRow):
                header_value = headers[col_index] if col_index < len(headers) else ""
                if 'footer' not in header_value.lower():
                    clean_value = smart_convert(cell_value)

            index += 1
            try:
                taskNum = labourRow[col_map['detail_WBS2']]
                subtaskNum = labourRow[col_map['detail_WBS3']]
                date = labourRow[col_map['detail_transDate']]
                employeeID = labourRow[col_map['detail_Employee']]
                billingStatus = labourRow[col_map['detail_billStatus']].strip()
                hours = float(labourRow[col_map['detail_totHrs']] or 0)

                if (billingStatus == "T"):
                    transfer_key = f'{employeeID}#{taskNum}#{subtaskNum}#{date}'
                    hours = transfer_map.get(transfer_key)
                    chargeHours = f"{hours:.2f}"
                    key = f'{employeeID}#{taskNum}#{subtaskNum}#{date}#F#{chargeHours}'

                    if key in invoice_map:
                        for i in invoice_map[key]:
                            invoice_row = int(i.split("#")[0])
                            invoice_number = f"{i.split('#')[1]}"
                            ws.cell(row = index-1, column=col_map['Comment']+1).value = invoice_number
                            break
            except Exception as e:
                continue
    
    ws.cell(row=1, column=col_map['Comment']+1).value = "invoice"
    transactionFileName = "Labour_Details_With_Invoices " + projectName + ".xlsx"
    CONSOLE_OUTPUT.tqdm_write(f"✅ {projectName} timesheet invoice match finished")
    CONSOLE_OUTPUT.tqdm_write(f"---------------------------------")
    wb.save( os.path.join("bridge_report", transactionFileName))

def labour_merge_transaction_details():
    folder_path = os.path.join("bridge_report")
    output_file = "merged_Transation Detail.csv"
    all_files = sorted([
        f for f in os.listdir(folder_path)
        if f.startswith("Transaction Detail ") and f.endswith(".csv")
    ])

    output_path = os.path.join(folder_path,"output" ,output_file)

    with open(output_path, 'w', newline='', encoding='utf-8-sig') as fout:
        writer = None
        
        for idx, file_name in enumerate(all_files):
            file_path = os.path.join(folder_path, file_name)
            with open(file_path, 'r', encoding='utf-8-sig') as fin:
                reader = csv.reader(fin)
                rows = list(reader)
                if not rows:
                    continue 
                if idx == 0:
                    writer = csv.writer(fout)
                    writer.writerows(rows)
                else:
                    writer.writerows(rows[1:-1])

def labour_merge_labour_details():
    folder_path = os.path.join("bridge_report")
    output_file = "merged_Labour_Details_With_Invoices.csv"

    all_files = sorted([
        f for f in os.listdir(folder_path)
        if f.startswith("Labour_Details_With_Invoices") and f.endswith(".xlsx")
    ])
    output_path = os.path.join(folder_path, "output", output_file)
    with open(output_path, 'w', newline='', encoding='utf-8-sig') as fout:
        writer = csv.writer(fout)
        header_written = False
        for file_name in all_files:
            file_path = os.path.join(folder_path, file_name)
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            if not header_written:
                writer.writerow(rows[0])
                header_written = True
            for row in rows[1:]:
                if row is None:
                    continue
                if all(cell is None or str(cell).strip() == "" for cell in row):  
                    continue
                writer.writerow(row)

def main():
    util.check_folder("bridge_report")
    util.check_folder("bridge_report/output")
    util.clear_folder("bridge_report")
    download_invoice_register_ytd()
    projectList = set()
    with open('bridge_report/Invoice Register.csv', newline='', encoding='utf-8') as csvfile:
        reader = list(csv.reader(csvfile))
        header = reader[3]
        col_map = {col_name.strip(): idx for idx, col_name in enumerate(header)}
        for index, invoiceRow in enumerate(reader):
            if index < 4:
                continue
            if not invoiceRow or len(invoiceRow) <= 13:
                continue
            projectId = invoiceRow[col_map['detail_WBS1']]
            # invoiceId = invoiceRow[13]
            if not projectId in projectList:
                projectList.add(projectId)
                labour_download_labour_details(projectId)
                labour_download_transaction_details(projectId)
                labour_invoice_match(projectId)
                
    labour_merge_transaction_details()
    labour_merge_labour_details()


if __name__ == "__main__":
    main()